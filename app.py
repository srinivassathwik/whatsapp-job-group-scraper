"""
C2C WhatsApp Job Scraper — Unified App
=======================================
ONE FILE. Combines config editor + scraper control + job dashboard.

Install:  pip install flask playwright && playwright install chromium
Run:      python app.py
Open:     http://127.0.0.1:5000
"""

import json, re, subprocess, sys, threading, time
from datetime import datetime
from pathlib import Path
from flask import Flask, jsonify, Response, request

app   = Flask(__name__)
SDIR  = Path(__file__).resolve().parent
CFG   = SDIR / "config.json"
OUT   = SDIR / "output"
OUT.mkdir(exist_ok=True)

# ── scraper state (shared across threads) ────────────────────────────────
scraper_state = {
    "running": False,
    "log":     [],
    "started": None,
    "finished": None,
    "result":  None,   # "ok" | "error"
}
scraper_lock = threading.Lock()

# ── helpers ──────────────────────────────────────────────────────────────
def read_cfg():
    if not CFG.exists():
        return {}
    with open(CFG, encoding="utf-8") as f:
        raw = json.load(f)
    return {k: v for k, v in raw.items()
            if not k.startswith("-") and not k.startswith("_")
            and k not in ("────── GROUPS ──────────────────────────────────────────────",)}

def write_cfg(data: dict):
    # Preserve comment keys if file already exists
    existing = {}
    if CFG.exists():
        with open(CFG, encoding="utf-8") as f:
            existing = json.load(f)
    # Merge: update only known editable keys
    for k, v in data.items():
        existing[k] = v
    with open(CFG, "w", encoding="utf-8") as f:
        json.dump(existing, f, indent=2, ensure_ascii=False)

def load_json(fname):
    p = OUT / fname
    if not p.exists():
        return []
    with open(p, encoding="utf-8") as f:
        return json.load(f)

EXP_BUCKETS = {"0-3":(0,3),"3-6":(3,6),"6-9":(6,9),"9-12":(9,12),"12+":(12,999)}

def parse_exp(s):
    if not s: return None
    m = re.search(r"(\d+)\s*[-–to]+\s*(\d+)", s)
    if m: return (int(m.group(1)), int(m.group(2)))
    m = re.search(r"(\d+)\+", s)
    if m: return (int(m.group(1)), 99)
    m = re.search(r"(\d+)", s)
    if m: n=int(m.group(1)); return (n,n)

def exp_ok(exp_str, label):
    r = parse_exp(exp_str)
    if not r: return False
    bmin, bmax = EXP_BUCKETS.get(label,(0,0))
    return r[0]<=bmax and r[1]>=bmin

# ── scraper runner ────────────────────────────────────────────────────────
def run_scraper():
    global scraper_state
    with scraper_lock:
        scraper_state["running"]  = True
        scraper_state["log"]      = []
        scraper_state["started"]  = datetime.now().isoformat()
        scraper_state["finished"] = None
        scraper_state["result"]   = None

    def emit(line):
        with scraper_lock:
            scraper_state["log"].append(line)

    try:
        main_py = SDIR / "main.py"
        if not main_py.exists():
            emit("ERROR: main.py not found next to app.py")
            with scraper_lock:
                scraper_state["running"]  = False
                scraper_state["result"]   = "error"
                scraper_state["finished"] = datetime.now().isoformat()
            return

        proc = subprocess.Popen(
            [sys.executable, str(main_py)],
            stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, cwd=str(SDIR), bufsize=1
        )
        for line in proc.stdout:
            emit(line.rstrip())
        proc.wait()
        result = "ok" if proc.returncode == 0 else "error"
    except Exception as e:
        emit(f"ERROR: {e}")
        result = "error"

    with scraper_lock:
        scraper_state["running"]  = False
        scraper_state["result"]   = result
        scraper_state["finished"] = datetime.now().isoformat()

# ── API routes ────────────────────────────────────────────────────────────

@app.route("/api/config", methods=["GET"])
def api_cfg_get():
    return jsonify(read_cfg())

@app.route("/api/config", methods=["POST"])
def api_cfg_post():
    data = request.json
    write_cfg(data)
    return jsonify({"ok": True})

@app.route("/api/scrape/start", methods=["POST"])
def api_scrape_start():
    with scraper_lock:
        if scraper_state["running"]:
            return jsonify({"ok": False, "msg": "Already running"})
    t = threading.Thread(target=run_scraper, daemon=True)
    t.start()
    return jsonify({"ok": True})

@app.route("/api/scrape/status")
def api_scrape_status():
    with scraper_lock:
        return jsonify(dict(scraper_state))

@app.route("/api/jobs")
def api_jobs():
    jobs = load_json("jobs.json")
    q        = request.args.get("q","").strip().lower()
    title_q  = request.args.get("title","").strip().lower()
    skill_q  = request.args.get("skill","").strip().lower()
    contract = request.args.get("contract","").strip()
    visa     = request.args.get("visa","").strip()
    location = request.args.get("location","").strip().lower()
    group    = request.args.get("group","").strip()
    exp_b    = [b for b in request.args.get("experience","").split(",") if b.strip()]
    has_rate = request.args.get("has_rate","")
    has_mail = request.args.get("has_email","")

    def ok(j):
        if title_q:
            if title_q not in (j.get("job_title") or "").lower():
                return False
        if skill_q:
            skills_str = " ".join(j.get("skills") or []).lower()
            raw = (j.get("raw_message") or "").lower()
            if skill_q not in skills_str and skill_q not in raw:
                return False
        if q:
            hay=" ".join(filter(None,[j.get("job_title"),j.get("location"),
                j.get("client")," ".join(j.get("skills") or []),j.get("raw_message","")])).lower()
            if q not in hay: return False
        if contract and (j.get("contract_type") or "").upper()!=contract.upper(): return False
        if visa:
            jv=[v.upper().replace(" ","") for v in (j.get("visa_types") or [])]
            if visa.upper().replace(" ","") not in jv: return False
        if location and location not in (j.get("location") or "").lower(): return False
        if group and j.get("source_group")!=group: return False
        if exp_b and not any(exp_ok(j.get("experience"),b) for b in exp_b): return False
        if has_rate=="1" and not j.get("rate"): return False
        if has_mail=="1" and not j.get("contact_email"): return False
        return True

    filtered=[j for j in jobs if ok(j)]
    filtered.sort(key=lambda j:j.get("wa_timestamp") or j.get("scraped_at") or "",reverse=True)
    return jsonify({"total":len(jobs),"filtered":len(filtered),"jobs":filtered})

@app.route("/api/review")
def api_review():
    data = load_json("review.json")
    status = request.args.get("status","")
    if status:
        data=[r for r in data if r.get("status")==status]
    data.sort(key=lambda r:r.get("wa_timestamp") or "",reverse=True)
    return jsonify({"total":len(load_json("review.json")),"filtered":len(data),"data":data})

@app.route("/api/meta")
def api_meta():
    jobs = load_json("jobs.json")
    return jsonify({
        "contracts": sorted({j.get("contract_type") for j in jobs if j.get("contract_type")}),
        "visas":     sorted({v for j in jobs for v in (j.get("visa_types") or [])}),
        "groups":    sorted({j.get("source_group") for j in jobs if j.get("source_group")}),
        "exp_buckets": list(EXP_BUCKETS.keys()),
        "total":     len(jobs),
    })

@app.route("/api/stats")
def api_stats():
    jobs    = load_json("jobs.json")
    review  = load_json("review.json")
    raw     = load_json("raw_messages.json")
    return jsonify({
        "jobs":        len(jobs),
        "messages":    len(raw),
        "spam":        sum(1 for r in review if r.get("status")=="spam"),
        "no_keywords": sum(1 for r in review if r.get("status")=="no_keywords"),
        "last_scrape": scraper_state.get("finished"),
    })

def normalize_rate(rate_str):
    """
    Normalize rate strings to a numeric $/hr value.
    '$55/hr', '$55 per hour', '55-65/hr', 'Up to $70/hr' → float or None
    Returns (display_str, numeric_low, numeric_high)
    """
    if not rate_str:
        return rate_str, None, None
    s = rate_str.strip()
    # Extract all dollar amounts
    nums = re.findall(r'\$?\s*(\d+(?:\.\d+)?)', s)
    nums = [float(n) for n in nums if 10 <= float(n) <= 500]  # sane $/hr range
    if not nums:
        return rate_str, None, None
    low  = min(nums)
    high = max(nums)
    if low == high:
        display = f"${low:.0f}/hr"
    else:
        display = f"${low:.0f}–${high:.0f}/hr"
    return display, low, high

@app.route("/api/export/excel")
def api_export_excel():
    """Export jobs.json as a downloadable Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    jobs = load_json("jobs.json")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "C2C Jobs"

    # ── Header row styling ────────────────────────────────────────────
    HEADER_FILL = PatternFill("solid", fgColor="1E1A16")
    HEADER_FONT = Font(bold=True, color="F5A623", size=10)
    ALT_FILL    = PatternFill("solid", fgColor="252018")
    BORDER_SIDE = Side(style="thin", color="3A3228")
    CELL_BORDER = Border(
        bottom=Border(bottom=BORDER_SIDE).bottom
    )

    headers = [
        ("ID",            8),  ("Job Title",     40), ("Location",     25),
        ("Experience",   14),  ("Rate",          14), ("Rate (Low $/hr)", 14),
        ("Rate (High $/hr)", 14), ("Duration",  18), ("Contract",      12),
        ("Visa Types",   22),  ("Skills",        35), ("Client",        20),
        ("Contact Email",28),  ("Contact Phone", 18), ("Apply Link",    35),
        ("Source Group", 22),  ("Posted",        18), ("Sender",        20),
        ("Raw Message",  60),
    ]

    for col, (title, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # ── Data rows ─────────────────────────────────────────────────────
    for row_idx, job in enumerate(jobs, 2):
        rate_display, rate_low, rate_high = normalize_rate(job.get("rate"))
        fill = ALT_FILL if row_idx % 2 == 0 else None

        row_data = [
            job.get("id", ""),
            job.get("job_title", ""),
            job.get("location", ""),
            job.get("experience", ""),
            rate_display or "",
            rate_low or "",
            rate_high or "",
            job.get("duration", ""),
            job.get("contract_type", ""),
            ", ".join(job.get("visa_types") or []),
            ", ".join(job.get("skills") or []),
            job.get("client", ""),
            job.get("contact_email", ""),
            job.get("contact_phone", ""),
            job.get("apply_link", ""),
            job.get("source_group", ""),
            job.get("wa_timestamp", ""),
            job.get("sender", ""),
            (job.get("raw_message") or "")[:500],  # truncate long messages
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=(col == len(row_data)))
            if fill:
                cell.fill = fill
            if col == 13 and value:  # email column — make it a mailto link
                cell.hyperlink = f"mailto:{value}"
                cell.font = Font(color="7EB8E8", underline="single")
            if col == 15 and value and value.startswith("http"):
                cell.hyperlink = value
                cell.font = Font(color="7EB8E8", underline="single")

    # ── Auto-filter ───────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(jobs)+1}"

    # ── Sheet metadata ────────────────────────────────────────────────
    wb.properties.title   = "C2C WhatsApp Jobs"
    wb.properties.creator = "C2C Scraper"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"c2c_jobs_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.route("/api/data/sizes")
def api_data_sizes():
    """Return record counts and file sizes for all output files."""
    files = {
        "jobs":         "jobs.json",
        "raw":          "raw_messages.json",
        "review":       "review.json",
        "seen":         "seen_jobs.json",
    }
    result = {}
    for key, fname in files.items():
        p = OUT / fname
        if p.exists():
            try:
                data = json.loads(p.read_text(encoding="utf-8"))
                # seen_jobs.json has a different structure
                if isinstance(data, dict):
                    count = data.get("total", len(data.get("seen_ids", [])))
                else:
                    count = len(data)
                size_kb = round(p.stat().st_size / 1024, 1)
            except Exception:
                count = 0
                size_kb = round(p.stat().st_size / 1024, 1)
            result[key] = {"count": count, "size_kb": size_kb, "exists": True}
        else:
            result[key] = {"count": 0, "size_kb": 0, "exists": False}
    return jsonify(result)

@app.route("/api/data/delete", methods=["POST"])
def api_data_delete():
    """
    Delete (reset) one or more output files.
    Body: { "files": ["jobs", "raw", "review", "seen", "all"] }
    """
    if scraper_state["running"]:
        return jsonify({"ok": False, "msg": "Cannot delete while scraper is running."})

    FILE_MAP = {
        "jobs":   OUT / "jobs.json",
        "raw":    OUT / "raw_messages.json",
        "review": OUT / "review.json",
        "seen":   OUT / "seen_jobs.json",
    }
    EMPTY = {
        "jobs":   "[]",
        "raw":    "[]",
        "review": "[]",
        "seen":   '{"description":"Reset","total":0,"last_updated":"","seen_ids":[]}',
    }

    targets = request.json.get("files", [])
    if "all" in targets:
        targets = list(FILE_MAP.keys())

    deleted = []
    for key in targets:
        if key in FILE_MAP:
            p = FILE_MAP[key]
            # Write empty instead of deleting so scraper doesn't crash
            p.write_text(EMPTY[key], encoding="utf-8")
            deleted.append(key)

    return jsonify({"ok": True, "deleted": deleted})

@app.route("/")
def index():
    return Response(HTML, mimetype="text/html")

# ── Embedded HTML (single-page app) ──────────────────────────────────────
HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>C2C Job Scraper</title>
<style>
/* ── WARM AMBER / CHARCOAL THEME ────────────────────────────────────────── */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');
:root{
  /* Backgrounds — warm charcoal, not cold blue-black */
  --bg:      #141210;   /* very dark warm black */
  --panel:   #1e1a16;   /* dark warm brown-black */
  --panel2:  #252018;   /* slightly lighter panel */
  --border:  #3a3228;   /* warm brown border */

  /* Text */
  --text:    #f0e8d8;   /* warm off-white, like paper */
  --dim:     #8a7d6a;   /* warm mid-brown for secondary text */

  /* Accent — amber/gold as primary */
  --accent:      #f5a623;   /* rich amber */
  --accent-dim:  #3d2a0a;   /* dark amber for backgrounds */
  --accent-glow: #f5a62340; /* amber glow */

  /* Status colors — warm variants */
  --green:   #6fcf7a;   /* softer green */
  --green-dim:#1a2e1c;
  --blue:    #7eb8e8;   /* warm steel blue */
  --orange:  #f5a623;   /* same as accent */
  --purple:  #c4a0e8;   /* lavender */
  --red:     #e87070;   /* warm red */
  --yellow:  #f5c842;   /* bright gold */
  --cyan:    #7dd3c0;   /* warm teal */

  --mono:"JetBrains Mono","SF Mono",Consolas,monospace;
  --sans:"Inter",-apple-system,"Segoe UI",sans-serif;

  /* Heights for sticky offset calculation */
  --nav-h:   52px;
  --bar-h:   38px;
  --tab-h:   45px;
}
/* Light mode */
body.light{
  --bg:#f5f0e8;--panel:#ffffff;--panel2:#f0ebe0;--border:#d4c8b0;
  --text:#2a2018;--dim:#7a6a50;
  --accent:#c47d0a;--accent-dim:#fdf0d8;--accent-glow:#c47d0a30;
  --green:#2a7a3a;--green-dim:#e8f5eb;
  --blue:#1a5fa8;--orange:#c47d0a;--purple:#6a3a9a;
  --red:#b83a2a;--yellow:#8a5a00;--cyan:#1a6a5a;
}
body.light .log-box{background:#f8f4ee}
body.light #modal{background:#ffffff}
body.light .raw-box{background:#f8f4ee}
body.light tbody tr:hover{background:#fdf8f0}
body.light .tc2c{background:#fff8e8;border-color:#d4a830}
body.light .tw2{background:#e8f0ff;border-color:#4a7ac8}
body.light .tskill{background:#e8f8f5;border-color:#2a8a7a}
body.light .tok{background:#e8f5eb;border-color:#2a7a3a}
body.light .tspam{background:#fde8e8;border-color:#c84a3a}
*{box-sizing:border-box;margin:0;padding:0}
body{
  background:var(--bg);
  color:var(--text);
  font-family:var(--sans);
  font-size:13px;
  line-height:1.5;
  min-height:100vh;
  /* subtle warm texture */
  background-image: radial-gradient(ellipse at 20% 0%, #2a1f0a18 0%, transparent 60%),
                    radial-gradient(ellipse at 80% 100%, #1a0a0518 0%, transparent 60%);
}

/* ── Nav ── */
nav{
  background:var(--panel);
  border-bottom:1px solid var(--border);
  display:flex;align-items:center;padding:0 20px;
  height:var(--nav-h);gap:4px;
  position:sticky;top:0;z-index:100;
  box-shadow:0 1px 0 var(--border), 0 4px 12px #00000040;
}
.nav-logo{
  font-weight:700;font-size:15px;
  color:var(--accent);
  margin-right:16px;white-space:nowrap;
  letter-spacing:-.02em;
}
.nav-logo span{color:var(--dim);font-weight:400}
.nav-btn{
  background:none;border:none;color:var(--dim);
  padding:8px 14px;border-radius:6px;cursor:pointer;
  font-size:13px;font-family:var(--sans);transition:all .15s;white-space:nowrap;
}
.nav-btn:hover{background:var(--panel2);color:var(--text)}
.nav-btn.active{
  background:var(--accent-dim);
  color:var(--accent);
  font-weight:600;
  box-shadow:0 0 0 1px var(--accent)30;
}
.nav-stats{margin-left:auto;display:flex;gap:16px;font-family:var(--mono);font-size:11px;color:var(--dim)}
.nav-stats b{color:var(--accent)}

/* ── Pages ── */
.page{display:none;padding:24px;max-width:1400px;margin:0 auto}
.page.active{display:block}

/* ── Config page ── */
.cfg-grid{display:grid;grid-template-columns:1fr 1fr;gap:20px}
@media(max-width:900px){.cfg-grid{grid-template-columns:1fr}}
.cfg-card{
  background:var(--panel);
  border:1px solid var(--border);
  border-radius:10px;padding:20px;
  transition:border-color .2s;
}
.cfg-card:hover{border-color:#5a4a35}
.cfg-card h3{
  font-size:11px;font-weight:700;
  color:var(--accent);
  text-transform:uppercase;letter-spacing:.08em;
  margin-bottom:16px;
  display:flex;align-items:center;gap:8px;
  padding-bottom:10px;
  border-bottom:1px solid var(--border);
}
.cfg-card h3 .icon{font-size:16px}
.field{margin-bottom:14px}
.field label{display:block;font-size:11px;text-transform:uppercase;letter-spacing:.06em;color:var(--dim);margin-bottom:5px}
.field input[type=text],.field input[type=number],.field select{
  width:100%;background:var(--panel2);border:1px solid var(--border);
  color:var(--text);padding:8px 10px;border-radius:6px;font-size:13px;
  font-family:var(--sans);transition:border-color .15s;
}
.field input:focus,.field select:focus{outline:none;border-color:var(--accent)}
.field .hint{font-size:11px;color:var(--dim);margin-top:4px}
.toggle-row{display:flex;align-items:center;gap:10px}
.toggle{width:40px;height:22px;background:var(--border);border-radius:11px;border:none;cursor:pointer;position:relative;transition:background .2s;flex-shrink:0}
.toggle.on{background:var(--accent)}
.toggle::after{content:'';position:absolute;width:16px;height:16px;background:#fff;border-radius:50%;top:3px;left:3px;transition:left .2s}
.toggle.on::after{left:21px}

/* Groups list */
.groups-list{display:flex;flex-direction:column;gap:8px;margin-bottom:12px}
.group-row{display:flex;align-items:center;gap:8px;background:var(--panel2);border:1px solid var(--border);border-radius:6px;padding:8px 10px;transition:border-color .15s}
.group-row:hover{border-color:#5a4a35}
.group-row input{flex:1;background:none;border:none;color:var(--text);font-size:13px;font-family:var(--mono);outline:none}
.group-row .del-btn{background:none;border:none;color:var(--dim);cursor:pointer;font-size:16px;padding:0 4px;line-height:1;transition:color .15s}
.group-row .del-btn:hover{color:var(--red)}
.add-btn{background:none;border:1px dashed var(--border);color:var(--dim);padding:8px 14px;border-radius:6px;cursor:pointer;width:100%;font-size:12px;transition:all .15s}
.add-btn:hover{border-color:var(--accent);color:var(--accent)}
.save-btn{
  background:var(--accent);color:#141210;border:none;
  padding:10px 24px;border-radius:7px;font-weight:700;font-size:13px;
  cursor:pointer;transition:all .15s;margin-top:4px;
  box-shadow:0 2px 8px var(--accent-glow);
}
.save-btn:hover{opacity:.85;box-shadow:0 4px 16px var(--accent-glow)}
.save-msg{font-size:12px;color:var(--accent);font-family:var(--mono);display:none;margin-left:12px}
.save-msg.show{display:inline}

/* File cards */
.file-card{background:var(--panel2);border:1px solid var(--border);border-radius:8px;padding:14px;display:flex;flex-direction:column;gap:6px;transition:border-color .15s}
.file-card:hover{border-color:#5a4a35}
.fc-name{font-family:var(--mono);font-size:11px;color:var(--blue);font-weight:600}
.fc-stat{font-family:var(--mono);font-size:20px;font-weight:700;color:var(--text)}
.fc-desc{font-size:11px;color:var(--dim);flex:1}
.fc-del{background:none;border:1px solid var(--border);color:var(--dim);padding:5px 10px;border-radius:5px;cursor:pointer;font-size:11px;margin-top:4px;transition:all .15s;width:100%}
.fc-del:hover{border-color:var(--red);color:var(--red);background:#3a1a1a}
.del-all-btn{background:#3a1a1a;border:1px solid var(--red);color:var(--red);padding:10px 20px;border-radius:7px;font-weight:700;font-size:13px;cursor:pointer;transition:opacity .15s;white-space:nowrap}
.del-all-btn:hover{opacity:.8}

/* Tag list editor */
.tag-editor{display:flex;flex-wrap:wrap;gap:6px;min-height:44px;background:var(--panel2);border:1px solid var(--border);border-radius:6px;padding:8px;margin-bottom:8px}
.tag-editor .te-tag{display:flex;align-items:center;gap:4px;background:var(--accent-dim);border:1px solid #5a3a10;border-radius:4px;padding:3px 8px;font-family:var(--mono);font-size:11px;color:var(--accent)}
.tag-editor .te-tag button{background:none;border:none;color:#8a6030;cursor:pointer;font-size:13px;line-height:1;padding:0 2px;transition:color .1s}
.tag-editor .te-tag button:hover{color:var(--red)}
.te-input-row{display:flex;gap:6px}
.te-input-row input{flex:1;background:var(--panel2);border:1px solid var(--border);color:var(--text);padding:6px 10px;border-radius:5px;font-size:12px;font-family:var(--mono)}
.te-input-row input:focus{outline:none;border-color:var(--accent)}
.te-add-btn{background:var(--accent-dim);border:1px solid var(--accent);color:var(--accent);padding:6px 14px;border-radius:5px;font-size:12px;cursor:pointer;white-space:nowrap;transition:all .15s}
.te-add-btn:hover{background:var(--accent);color:#141210}

/* ── Scrape page ── */
.scrape-layout{display:grid;grid-template-columns:320px 1fr;gap:20px}
@media(max-width:900px){.scrape-layout{grid-template-columns:1fr}}
.scrape-card{background:var(--panel);border:1px solid var(--border);border-radius:10px;padding:24px}
.scrape-card h3{font-size:11px;font-weight:700;color:var(--accent);text-transform:uppercase;letter-spacing:.08em;margin-bottom:16px;padding-bottom:10px;border-bottom:1px solid var(--border)}
.big-run-btn{
  width:100%;padding:16px;
  background:var(--accent);color:#141210;border:none;border-radius:8px;
  font-size:15px;font-weight:700;cursor:pointer;transition:all .2s;
  display:flex;align-items:center;justify-content:center;gap:10px;
  box-shadow:0 2px 12px var(--accent-glow);
}
.big-run-btn:hover{opacity:.85;box-shadow:0 4px 20px var(--accent-glow)}
.big-run-btn:disabled{background:var(--border);color:var(--dim);cursor:not-allowed;opacity:1;box-shadow:none}
.big-run-btn.running{background:var(--accent-dim);color:var(--accent);border:1px solid var(--accent);box-shadow:0 0 20px var(--accent-glow)}
.status-badge{display:inline-flex;align-items:center;gap:6px;padding:4px 10px;border-radius:20px;font-size:11px;font-family:var(--mono);margin-top:12px}
.status-badge.idle{background:var(--panel2);color:var(--dim)}
.status-badge.running{background:var(--accent-dim);color:var(--accent)}
.status-badge.ok{background:var(--green-dim);color:var(--green)}
.status-badge.error{background:#3a1a1a;color:var(--red)}
.pulse{display:inline-block;width:8px;height:8px;border-radius:50%;background:currentColor}
.pulse.anim{animation:pulse 1s infinite}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.3}}
.log-box{
  background:#0e0c0a;
  border:1px solid var(--border);border-radius:8px;
  padding:14px;height:500px;overflow-y:auto;
  font-family:var(--mono);font-size:11.5px;line-height:1.7;color:var(--dim);
}
.log-box .log-info{color:var(--dim)}
.log-box .log-ok{color:var(--green)}
.log-box .log-job{color:var(--accent)}
.log-box .log-spam{color:var(--red)}
.log-box .log-warn{color:var(--yellow)}
.log-box .log-err{color:var(--red);font-weight:600}
.log-box .log-step{color:var(--cyan)}
.scrape-summary{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-top:20px}
.sum-card{background:var(--panel2);border:1px solid var(--border);border-radius:8px;padding:14px;text-align:center}
.sum-card .val{font-size:28px;font-weight:700;font-family:var(--mono)}
.sum-card .lbl{font-size:11px;color:var(--dim);text-transform:uppercase;letter-spacing:.06em;margin-top:4px}

/* ── Jobs column header bar (separate from table, always visible) ── */
.jobs-col-header{
  display:flex;align-items:center;
  background:var(--panel);
  border-bottom:2px solid var(--border);
  padding:0;
  position:sticky;
  top:90px;   /* nav(52) + stats-bar(38) */
  z-index:5;
  flex-shrink:0;
  min-width:1000px;
}
.jobs-col-header > div{
  padding:10px 12px;
  font-size:10px;font-weight:700;
  text-transform:uppercase;letter-spacing:.08em;
  color:var(--accent);
  white-space:nowrap;
  flex-shrink:0;
}
.jobs-sidebar{
  width:220px;min-width:220px;
  background:var(--panel);
  border-right:1px solid var(--border);
  padding:16px 14px;
  position:sticky;top:52px;
  height:calc(100vh - 52px);
  overflow-y:auto;flex-shrink:0;
}
.fg{margin-bottom:16px}
.fg .lbl{font-size:10px;text-transform:uppercase;letter-spacing:.07em;color:var(--dim);display:block;margin-bottom:5px}
.fg input,.fg select{
  width:100%;background:var(--panel2);border:1px solid var(--border);
  color:var(--text);padding:7px 8px;border-radius:5px;font-size:12px;font-family:var(--sans);
  transition:border-color .15s;
}
.fg input:focus,.fg select:focus{outline:none;border-color:var(--accent)}
.checks{display:flex;flex-direction:column;gap:5px}
.checks label{display:flex;align-items:center;gap:7px;font-size:12px;color:var(--text);cursor:pointer}
.checks input[type=checkbox]{accent-color:var(--accent)}
.tog-row{display:flex;gap:6px;flex-wrap:wrap;margin-top:2px}
.tog-btn{background:var(--panel2);border:1px solid var(--border);color:var(--dim);padding:4px 9px;border-radius:4px;font-size:11px;cursor:pointer;transition:all .12s}
.tog-btn.active{border-color:var(--accent);color:var(--accent);background:var(--accent-dim)}
#jobs-reset{width:100%;background:transparent;border:1px solid var(--border);color:var(--dim);padding:7px;border-radius:5px;cursor:pointer;font-size:11px;text-transform:uppercase;letter-spacing:.04em;transition:all .12s;margin-top:4px}
#jobs-reset:hover{border-color:var(--accent);color:var(--accent)}

/* review tabs */
.tab-row{
  display:flex;gap:0;border-bottom:1px solid var(--border);
  background:var(--panel);padding:0 20px;
  position:sticky;top:var(--nav-h);z-index:5;
  height:var(--tab-h);
}
.tab{
  display:flex;align-items:center;
  padding:0 18px;border-bottom:2px solid transparent;
  cursor:pointer;font-size:12px;color:var(--dim);
  transition:all .15s;white-space:nowrap;
}
.tab:hover{color:var(--text)}
.tab.active{color:var(--accent);border-bottom-color:var(--accent)}

/* stats bar */
.stats-bar{
  background:var(--panel);
  border-bottom:1px solid var(--border);
  padding:0 20px;
  display:flex;align-items:center;gap:20px;
  font-family:var(--mono);font-size:11px;color:var(--dim);
  z-index:4;
  height:38px;
  flex-shrink:0;
}
.stats-bar b{color:var(--accent)}

/* table */
table{width:100%;border-collapse:collapse;min-width:1000px}
thead th{
  text-align:left;font-size:10px;text-transform:uppercase;
  letter-spacing:.08em;color:var(--accent);
  padding:10px 12px;border-bottom:2px solid var(--border);
  background:var(--panel);z-index:3;white-space:nowrap;
  font-weight:700;
}

tbody tr{border-bottom:1px solid var(--border);cursor:pointer;transition:background .1s}
tbody tr:hover{background:var(--panel2)}
tbody tr:hover td:first-child{border-left:2px solid var(--accent)}
td{padding:10px 12px;vertical-align:top}
.cell-title{font-weight:600;font-size:13px;line-height:1.4;word-break:break-word;color:var(--text)}
.cell-group{font-size:10px;color:var(--dim);font-family:var(--mono);margin-top:3px}
.cell-tags{display:flex;flex-wrap:wrap;gap:3px;margin-top:5px}
.tag{font-family:var(--mono);font-size:10px;padding:2px 7px;border-radius:3px;border:1px solid transparent;white-space:nowrap;font-weight:500}
.tc2c{color:#f5c842;border-color:#5a4010;background:#2a1e05}
.tw2{color:var(--blue);border-color:#1d3557;background:#0d1e30}
.tc2h{color:var(--orange);border-color:#4a2a0e;background:#200e00}
.tcon{color:var(--purple);border-color:#3a244a;background:#1a0e25}
.tvisa{color:var(--dim);border-color:var(--border)}
.tskill{color:var(--cyan);border-color:#1a3a35;background:#0a1e1a}
.tspam{color:var(--red);border-color:#3a1a1a;background:#200808}
.tok{color:var(--green);border-color:#1d4428;background:#0a1e0e}
.mono{font-family:var(--mono);font-size:11.5px;color:var(--dim);word-break:break-word}
.green-val{font-family:var(--mono);font-size:12px;color:var(--accent);font-weight:700}
.contact-cell{font-family:var(--mono);font-size:11px;word-break:break-all}
.contact-cell a{color:var(--blue);text-decoration:none}
.contact-cell a:hover{text-decoration:underline;color:var(--accent)}
.time-cell{font-family:var(--mono);font-size:11px;color:var(--dim);white-space:nowrap}
.empty{padding:60px;text-align:center;color:var(--dim);font-size:14px}
.tbl-wrap{overflow-x:auto}

/* modal */
#overlay{position:fixed;inset:0;background:rgba(0,0,0,.75);display:none;align-items:center;justify-content:center;z-index:200;backdrop-filter:blur(3px)}
#overlay.open{display:flex}
#modal{
  background:var(--panel);
  border:1px solid var(--border);
  border-radius:12px;width:92%;max-width:720px;max-height:88vh;
  overflow-y:auto;padding:26px 28px;position:relative;
  box-shadow:0 20px 60px #00000080, 0 0 0 1px var(--border);
}
.mclose{position:absolute;top:14px;right:18px;background:none;border:none;color:var(--dim);font-size:22px;cursor:pointer;transition:color .15s}
.mclose:hover{color:var(--accent)}
#modal h2{font-size:17px;font-weight:700;margin-bottom:8px;padding-right:30px;word-break:break-word;color:var(--text)}
.m-pills{display:flex;flex-wrap:wrap;gap:5px;margin-bottom:14px}
.m-grid{display:grid;grid-template-columns:1fr 1fr;gap:10px 20px;margin-bottom:16px}
@media(max-width:540px){.m-grid{grid-template-columns:1fr}}
.mf{border-bottom:1px solid var(--border);padding-bottom:8px}
.mf .ml{font-size:10px;text-transform:uppercase;letter-spacing:.07em;color:var(--dim);margin-bottom:3px}
.mf .mv{font-family:var(--mono);font-size:12px;color:var(--text);word-break:break-word}
.mf .mv a{color:var(--blue)}
.mf .mv.gv{color:var(--accent);font-weight:700}
.mf.full{grid-column:1/-1}
.raw-box{white-space:pre-wrap;font-size:12px;line-height:1.65;color:var(--dim);background:#0e0c0a;border:1px solid var(--border);border-radius:6px;padding:14px;margin-top:14px;font-family:var(--mono);max-height:260px;overflow-y:auto}
.raw-toggle{background:none;border:1px solid var(--border);color:var(--dim);padding:5px 12px;border-radius:5px;font-size:11px;cursor:pointer;margin-top:12px;transition:all .15s}
.raw-toggle:hover{border-color:var(--accent);color:var(--accent)}
</style>
</head>
<body>

<nav>
  <div class="nav-logo">◆ C2C Scraper <span>v2</span></div>
  <button class="nav-btn active" onclick="showPage('config',this)">⚙ Config</button>
  <button class="nav-btn" onclick="showPage('scrape',this)">▶ Scrape</button>
  <button class="nav-btn" onclick="showPage('jobs',this)">📋 Jobs</button>
  <button class="nav-btn" onclick="showPage('review',this)">🔍 Review</button>
  <div class="nav-stats" id="nav-stats">loading...</div>
  <button id="theme-btn" onclick="toggleTheme()" style="background:none;border:1px solid var(--border);color:var(--dim);padding:5px 12px;border-radius:6px;cursor:pointer;font-size:12px;margin-left:8px;transition:all .15s">🌙 Dark</button>
</nav>

<!-- ══════════════════════════════════════════════════════════════
     PAGE 1 — CONFIG
══════════════════════════════════════════════════════════════ -->
<div class="page active" id="page-config">
  <div style="display:flex;align-items:center;gap:14px;margin-bottom:22px">
    <h2 style="font-size:18px;font-weight:700">Configuration</h2>
    <button class="save-btn" onclick="saveConfig()">Save Config</button>
    <span class="save-msg" id="save-msg">Saved!</span>
  </div>

  <div class="cfg-grid">
    <!-- Groups -->
    <div class="cfg-card" style="grid-column:1/-1">
      <h3><span class="icon">👥</span> WhatsApp Groups to Scrape</h3>
      <div class="groups-list" id="groups-list"></div>
      <button class="add-btn" onclick="addGroup()">+ Add Group</button>
      <div style="margin-top:10px;font-size:11px;color:var(--dim)">
        Group names must exactly match WhatsApp (spaces, pipes, commas all matter)
      </div>
    </div>

    <!-- Scraping limits -->
    <div class="cfg-card">
      <h3><span class="icon">⏱</span> Scraping Limits</h3>
      <div class="field">
        <label>Max messages per group</label>
        <input type="number" id="max_messages_per_group" min="10" max="1000">
        <div class="hint">How far back to scroll in each group (default: 200)</div>
      </div>
      <div class="field">
        <label>Only last N hours</label>
        <input type="number" id="only_last_hours" min="1" max="168">
        <div class="hint">Only keep messages from last N hours (default: 24)</div>
      </div>
    </div>

    <!-- Browser settings -->
    <div class="cfg-card">
      <h3><span class="icon">🌐</span> Browser Settings</h3>
      <div class="field">
        <label>Login timeout (seconds)</label>
        <input type="number" id="login_timeout_seconds" min="30" max="300">
        <div class="hint">Time to wait for QR code scan (default: 120)</div>
      </div>
      <div class="field">
        <label>Headless mode</label>
        <div class="toggle-row">
          <button class="toggle" id="headless-toggle" onclick="toggleHeadless()"></button>
          <span id="headless-label" style="font-size:12px;color:var(--dim)">Off (browser visible)</span>
        </div>
        <div class="hint">On = run Chrome in background. Off = see browser window (needed for first QR scan)</div>
      </div>
    </div>

    <!-- Job Keywords editor -->
    <div class="cfg-card">
      <h3><span class="icon">🔑</span> Job Keywords</h3>
      <div class="tag-editor" id="kw-tags"></div>
      <div class="te-input-row">
        <input type="text" id="kw-input" placeholder="Add keyword…" onkeydown="if(event.key==='Enter')addTag('kw')">
        <button class="te-add-btn" onclick="addTag('kw')">+ Add</button>
      </div>
      <div style="font-size:11px;color:var(--dim);margin-top:8px">Any ONE keyword match = message treated as job post.</div>
    </div>

    <!-- Visa types editor -->
    <div class="cfg-card">
      <h3><span class="icon">🛂</span> Visa Types</h3>
      <div class="tag-editor" id="visa-tags"></div>
      <div class="te-input-row">
        <input type="text" id="visa-input" placeholder="Add visa type…" onkeydown="if(event.key==='Enter')addTag('visa')">
        <button class="te-add-btn" onclick="addTag('visa')">+ Add</button>
      </div>
      <div style="font-size:11px;color:var(--dim);margin-top:8px">Extracted into <code>visa_types</code> field in jobs.json.</div>
    </div>

    <!-- Data Management -->
    <div class="cfg-card" style="grid-column:1/-1">
      <h3><span class="icon">🗂</span> Data Management</h3>

      <!-- File info table -->
      <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:18px" id="file-cards">
        <div class="file-card" id="fc-jobs">
          <div class="fc-name">jobs.json</div>
          <div class="fc-stat" id="fstat-jobs">–</div>
          <div class="fc-desc">Extracted job posts</div>
          <button class="fc-del" onclick="deleteFile('jobs')">Reset</button>
        </div>
        <div class="file-card" id="fc-raw">
          <div class="fc-name">raw_messages.json</div>
          <div class="fc-stat" id="fstat-raw">–</div>
          <div class="fc-desc">All scraped messages</div>
          <button class="fc-del" onclick="deleteFile('raw')">Reset</button>
        </div>
        <div class="file-card" id="fc-review">
          <div class="fc-name">review.json</div>
          <div class="fc-stat" id="fstat-review">–</div>
          <div class="fc-desc">Messages with filter reason</div>
          <button class="fc-del" onclick="deleteFile('review')">Reset</button>
        </div>
        <div class="file-card" id="fc-seen">
          <div class="fc-name">seen_jobs.json</div>
          <div class="fc-stat" id="fstat-seen">–</div>
          <div class="fc-desc">Prevents re-scraping</div>
          <button class="fc-del" onclick="deleteFile('seen')">Reset</button>
        </div>
      </div>

      <div style="display:flex;align-items:center;gap:12px;padding-top:14px;border-top:1px solid var(--border)">
        <button class="del-all-btn" onclick="deleteFile('all')">⚠ Reset ALL Output Files</button>
        <span style="font-size:11px;color:var(--dim)">
          Resetting seen_jobs.json allows re-scraping already-seen messages.<br>
          Resetting jobs.json clears all saved job data. Cannot be undone.
        </span>
      </div>
      <div id="del-msg" style="display:none;font-size:12px;font-family:var(--mono);margin-top:10px"></div>
    </div>
  </div>
</div>

<!-- ══════════════════════════════════════════════════════════════
     PAGE 2 — SCRAPE
══════════════════════════════════════════════════════════════ -->
<div class="page" id="page-scrape">
  <div style="margin-bottom:22px">
    <h2 style="font-size:18px;font-weight:700">Run Scraper</h2>
  </div>

  <div class="scrape-layout">
    <div>
      <div class="scrape-card" style="margin-bottom:16px">
        <h3>Control</h3>
        <button class="big-run-btn" id="run-btn" onclick="startScrape()">
          <span id="run-icon">▶</span> <span id="run-label">Start Scraping</span>
        </button>
        <div style="margin-top:12px">
          <div id="status-badge" class="status-badge idle">
            <span class="pulse" id="pulse-dot"></span>
            <span id="status-text">Idle</span>
          </div>
        </div>
        <div id="run-times" style="font-family:var(--mono);font-size:11px;color:var(--dim);margin-top:10px;display:none">
          <div id="run-started"></div>
          <div id="run-finished"></div>
        </div>
      </div>

      <div class="scrape-card">
        <h3>Current Config</h3>
        <div id="scrape-cfg-preview" style="font-family:var(--mono);font-size:12px;color:var(--dim);line-height:1.9"></div>
      </div>

      <div class="scrape-summary">
        <div class="sum-card"><div class="val" id="s-jobs" style="color:var(--green)">–</div><div class="lbl">Jobs</div></div>
        <div class="sum-card"><div class="val" id="s-msgs" style="color:var(--blue)">–</div><div class="lbl">Messages</div></div>
        <div class="sum-card"><div class="val" id="s-spam" style="color:var(--orange)">–</div><div class="lbl">Spam</div></div>
        <div class="sum-card"><div class="val" id="s-nokw" style="color:var(--dim)">–</div><div class="lbl">No Keywords</div></div>
      </div>
    </div>

    <div class="scrape-card">
      <h3>Live Log</h3>
      <div class="log-box" id="log-box"><div style="color:var(--dim)">Waiting to start...</div></div>
    </div>
  </div>
</div>

<!-- ══════════════════════════════════════════════════════════════
     PAGE 3 — JOBS
══════════════════════════════════════════════════════════════ -->
<div class="page" id="page-jobs" style="padding:0;max-width:none">
  <div style="display:flex;min-height:calc(100vh - 52px)">
    <aside class="jobs-sidebar">
      <div class="fg">
        <span class="lbl">Title Search</span>
        <input id="jtitle" type="text" placeholder="Java Developer, SAP Lead…">
      </div>
      <div class="fg">
        <span class="lbl">Skill Search</span>
        <input id="jskill" type="text" placeholder="Java, React, AWS, Python…">
      </div>
      <div class="fg">
        <span class="lbl">Contract Type</span>
        <select id="jcontract"><option value="">All types</option></select>
      </div>
      <div class="fg">
        <span class="lbl">Visa / Work Auth</span>
        <select id="jvisa"><option value="">All visas</option></select>
      </div>
      <div class="fg">
        <span class="lbl">Source Group</span>
        <select id="jgroup"><option value="">All groups</option></select>
      </div>
      <div class="fg">
        <span class="lbl">Experience (yrs)</span>
        <div id="jexp-checks" class="checks"></div>
      </div>
      <div class="fg">
        <span class="lbl">Location contains</span>
        <input id="jloc" type="text" placeholder="TX, Remote, NY…">
      </div>
      <div class="fg">
        <span class="lbl">Quick Filters</span>
        <div class="tog-row">
          <button class="tog-btn" data-f="has_rate" onclick="toggleTog(this)">$ Rate</button>
          <button class="tog-btn" data-f="has_email" onclick="toggleTog(this)">Email</button>
        </div>
      </div>
      <button id="jobs-reset" onclick="resetJobFilters()">↺ Reset</button>
    </aside>

    <div style="flex:1;min-width:0;display:flex;flex-direction:column">
      <!-- Stats bar: sticks below nav -->
      <div class="stats-bar" id="jobs-stats" style="position:sticky;top:52px;z-index:6;flex-shrink:0">Loading…
        <button onclick="exportExcel()" style="margin-left:auto;background:var(--accent);color:#141210;border:none;padding:5px 14px;border-radius:5px;font-size:11px;font-weight:700;cursor:pointer;font-family:var(--sans)">⬇ Export Excel</button>
      </div>
      <!-- Column headers: sticky below stats bar, always visible -->
      <div class="jobs-col-header" id="jobs-col-header">
        <div style="width:22%">Job Title</div>
        <div style="width:15%">Location</div>
        <div style="width:9%">Experience</div>
        <div style="width:9%">Rate</div>
        <div style="width:10%">Duration</div>
        <div style="width:11%">Visa / Contract</div>
        <div style="width:14%">Contact</div>
        <div style="width:10%">Posted</div>
      </div>
      <!-- Scrollable table body -->
      <div class="tbl-wrap">
        <table>
          <colgroup>
            <col style="width:22%"><col style="width:15%"><col style="width:9%">
            <col style="width:9%"><col style="width:10%"><col style="width:11%">
            <col style="width:14%"><col style="width:10%">
          </colgroup>
          <tbody id="jobs-rows"></tbody>
        </table>
        <div id="jobs-empty" class="empty" style="display:none">No jobs match your filters.</div>
      </div>
    </div>
  </div>
</div>

<!-- ══════════════════════════════════════════════════════════════
     PAGE 4 — REVIEW
══════════════════════════════════════════════════════════════ -->
<div class="page" id="page-review" style="padding:0;max-width:none">
  <div class="tab-row">
    <div class="tab active" onclick="setReviewTab('all',this)">All Messages</div>
    <div class="tab" onclick="setReviewTab('job',this)">✅ Jobs</div>
    <div class="tab" onclick="setReviewTab('spam',this)">🚫 Spam</div>
    <div class="tab" onclick="setReviewTab('no_keywords',this)">⚠ No Keywords</div>
  </div>
  <div class="stats-bar" id="review-stats" style="position:sticky;top:97px;z-index:4">Loading…</div>
  <div class="tbl-wrap">
    <table>
      <thead><tr>
        <th style="width:10%;position:sticky;top:135px;background:var(--panel);z-index:3">Status</th>
        <th style="width:12%;position:sticky;top:135px;background:var(--panel);z-index:3">Group</th>
        <th style="width:12%;position:sticky;top:135px;background:var(--panel);z-index:3">Posted</th>
        <th style="width:20%;position:sticky;top:135px;background:var(--panel);z-index:3">Filter Reason</th>
        <th style="width:46%;position:sticky;top:135px;background:var(--panel);z-index:3">Message Preview</th>
      </tr></thead>
      <tbody id="review-rows"></tbody>
    </table>
    <div id="review-empty" class="empty" style="display:none">No messages.</div>
  </div>
</div>

<!-- Modal -->
<div id="overlay">
  <div id="modal">
    <button class="mclose" onclick="closeModal()">×</button>
    <h2 id="m-title"></h2>
    <div class="m-pills" id="m-pills"></div>
    <div class="m-grid" id="m-grid"></div>
    <div style="display:flex;align-items:center;gap:8px;margin-top:12px">
      <button class="raw-toggle" id="raw-toggle" onclick="toggleRaw()">Show raw message</button>
      <button id="copy-raw-btn" onclick="copyRaw()" style="display:none;background:none;border:1px solid var(--border);color:var(--dim);padding:5px 12px;border-radius:5px;font-size:11px;cursor:pointer;transition:all .15s">📋 Copy</button>
    </div>
    <pre class="raw-box" id="m-raw" style="display:none;white-space:pre-wrap;word-break:break-word"></pre>
  </div>
</div>

<script>
// ── Helpers ────────────────────────────────────────────────────────────────
const $=id=>document.getElementById(id)
function esc(s){const d=document.createElement('div');d.textContent=s??'';return d.innerHTML}
function tag(t,cls){return`<span class="tag ${cls}">${esc(t)}</span>`}
function ctag(ct){
  if(!ct)return'';const u=ct.toUpperCase();
  if(u.includes('C2C')||u.includes('CORP'))return tag(ct,'tc2c');
  if(u==='W2')return tag(ct,'tw2');
  if(u.includes('C2H')||u.includes('HIRE'))return tag(ct,'tc2h');
  return tag(ct,'tcon');
}
function timeAgo(ts){
  if(!ts)return'–';
  let d;
  const m=ts.match(/(\d+):(\d+)\s*(AM|PM),\s*(\d+)\/(\d+)\/(\d+)/i);
  if(m){
    let h=+m[1],mn=+m[2];const ap=m[3].toUpperCase(),mo=m[4],dy=m[5],yr=m[6];
    if(ap==='PM'&&h<12)h+=12;if(ap==='AM'&&h===12)h=0;
    d=new Date(`${yr}-${mo.padStart(2,'0')}-${dy.padStart(2,'0')}T${String(h).padStart(2,'0')}:${String(mn).padStart(2,'0')}:00`);
  }else d=new Date(ts);
  if(isNaN(d))return ts;
  const hrs=Math.floor((Date.now()-d)/3600000);
  if(hrs<1)return'just now';if(hrs<24)return hrs+'h ago';
  return Math.floor(hrs/24)+'d ago';
}

// ── Nav ────────────────────────────────────────────────────────────────────
function showPage(name,btn){
  document.querySelectorAll('.page').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.nav-btn').forEach(b=>b.classList.remove('active'));
  $('page-'+name).classList.add('active');
  btn.classList.add('active');
  if(name==='jobs'){loadJobMeta();loadJobs();}
  if(name==='review'){loadReview();}
  if(name==='scrape'){loadScrapeCfgPreview();loadStats();}
  if(name==='config'){loadConfig();loadFileSizes();}
}
function loadNavStats(){
  fetch('/api/stats').then(r=>r.json()).then(d=>{
    $('nav-stats').innerHTML=`<b>${d.jobs}</b> jobs &nbsp;|&nbsp; <b>${d.messages}</b> msgs`;
  });
}
loadNavStats();setInterval(loadNavStats,15000);

// ── File sizes ─────────────────────────────────────────────────────────────
function loadFileSizes(){
  fetch('/api/data/sizes').then(r=>r.json()).then(d=>{
    const labels={jobs:'jobs',raw:'messages',review:'messages',seen:'IDs'};
    for(const[k,v] of Object.entries(d)){
      const el=$('fstat-'+k);
      if(!el)continue;
      if(v.exists){
        el.innerHTML=`<span style="color:var(--green)">${v.count}</span> <span style="font-size:11px;color:var(--dim)">${labels[k]||''} &nbsp; ${v.size_kb}KB</span>`;
      }else{
        el.innerHTML=`<span style="color:var(--dim)">empty</span>`;
      }
    }
  });
}
loadFileSizes();

function deleteFile(key){
  const labels={
    jobs:'jobs.json (all job data)',
    raw:'raw_messages.json (all messages)',
    review:'review.json (review data)',
    seen:'seen_jobs.json (re-scrape everything next run)',
    all:'ALL output files'
  };
  const msg=`Reset ${labels[key]||key}?\n\nThis clears the file. It cannot be undone.`;
  if(!confirm(msg))return;
  fetch('/api/data/delete',{
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body:JSON.stringify({files:[key]})
  }).then(r=>r.json()).then(d=>{
    const dm=$('del-msg');
    if(d.ok){
      dm.style.display='block';
      dm.style.color='var(--green)';
      dm.textContent=`Reset: ${d.deleted.join(', ')} — done.`;
      loadFileSizes();loadNavStats();
    }else{
      dm.style.display='block';
      dm.style.color='var(--red)';
      dm.textContent=`Error: ${d.msg}`;
    }
    setTimeout(()=>{dm.style.display='none';},4000);
  });
}

// ══════════════════════════════════════════════════════════════
//  CONFIG
// ══════════════════════════════════════════════════════════════
let cfg={};
function renderTags(containerId, items){
  const el=$(containerId);el.innerHTML='';
  (items||[]).forEach(v=>{
    const span=document.createElement('span');span.className='te-tag';
    span.innerHTML=`${esc(v)}<button onclick="removeTag('${containerId}','${v.replace(/'/g,"\\'")}')">×</button>`;
    el.appendChild(span);
  });
}
function addTag(type){
  const inputId=type==='kw'?'kw-input':'visa-input';
  const containerId=type==='kw'?'kw-tags':'visa-tags';
  const val=$(inputId).value.trim();
  if(!val)return;
  // Get current tags
  const current=[...$( containerId).querySelectorAll('.te-tag')].map(t=>t.textContent.trim().slice(0,-1));
  if(current.includes(val)){$(inputId).value='';return;}
  renderTags(containerId,[...current,val]);
  $(inputId).value='';
}
function removeTag(containerId,val){
  const current=[...$( containerId).querySelectorAll('.te-tag')].map(t=>t.textContent.trim().slice(0,-1));
  renderTags(containerId,current.filter(v=>v!==val));
}
function getTagList(containerId){
  return[...$( containerId).querySelectorAll('.te-tag')].map(t=>t.textContent.trim().slice(0,-1));
}
function loadConfig(){
  fetch('/api/config').then(r=>r.json()).then(d=>{
    cfg=d;
    // Groups
    const gl=$('groups-list');gl.innerHTML='';
    (d.groups||[]).forEach(g=>addGroupRow(g));
    // Numbers
    $('max_messages_per_group').value=d.max_messages_per_group||200;
    $('only_last_hours').value=d.only_last_hours||24;
    $('login_timeout_seconds').value=d.login_timeout_seconds||120;
    // Headless
    const ht=$('headless-toggle');
    if(d.headless){ht.classList.add('on');}else{ht.classList.remove('on');}
    updateHeadlessLabel();
    // Tag editors
    renderTags('kw-tags', d.job_keywords||[]);
    renderTags('visa-tags', d.visa_types||[]);
  });
}
function addGroupRow(val=''){
  const row=document.createElement('div');row.className='group-row';
  row.innerHTML=`<input type="text" value="${esc(val)}" placeholder="Exact WhatsApp group name">
    <button class="del-btn" onclick="this.parentElement.remove()">×</button>`;
  $('groups-list').appendChild(row);
}
function addGroup(){addGroupRow('');$('groups-list').lastElementChild.querySelector('input').focus();}
function toggleHeadless(){
  const t=$('headless-toggle');t.classList.toggle('on');updateHeadlessLabel();
}
function updateHeadlessLabel(){
  const on=$('headless-toggle').classList.contains('on');
  $('headless-label').textContent=on?'On (background mode)':'Off (browser visible)';
}
function saveConfig(){
  const groups=[...$('groups-list').querySelectorAll('input')].map(i=>i.value.trim()).filter(Boolean);
  const data={
    groups,
    max_messages_per_group:+$('max_messages_per_group').value||200,
    only_last_hours:+$('only_last_hours').value||24,
    login_timeout_seconds:+$('login_timeout_seconds').value||120,
    headless:$('headless-toggle').classList.contains('on'),
    job_keywords: getTagList('kw-tags'),
    visa_types:   getTagList('visa-tags'),
  };
  fetch('/api/config',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(data)})
    .then(r=>r.json()).then(()=>{
      const m=$('save-msg');m.classList.add('show');
      setTimeout(()=>m.classList.remove('show'),2500);
    });
}
loadConfig();

// ══════════════════════════════════════════════════════════════
//  SCRAPE
// ══════════════════════════════════════════════════════════════
let pollTimer=null;
function loadScrapeCfgPreview(){
  fetch('/api/config').then(r=>r.json()).then(d=>{
    $('scrape-cfg-preview').innerHTML=[
      `<div style="color:var(--dim)">Groups:</div>`,
      ...(d.groups||[]).map(g=>`<div style="color:var(--text);margin-left:8px">• ${esc(g)}</div>`),
      `<div style="margin-top:8px;color:var(--dim)">Last <b style="color:var(--green)">${d.only_last_hours||24}h</b> &nbsp;|&nbsp; Max <b style="color:var(--green)">${d.max_messages_per_group||200}</b> msgs/group</div>`
    ].join('');
  });
}
function loadStats(){
  fetch('/api/stats').then(r=>r.json()).then(d=>{
    $('s-jobs').textContent=d.jobs;
    $('s-msgs').textContent=d.messages;
    $('s-spam').textContent=d.spam;
    $('s-nokw').textContent=d.no_keywords;
  });
}
function startScrape(){
  fetch('/api/scrape/start',{method:'POST'}).then(r=>r.json()).then(d=>{
    if(!d.ok){alert(d.msg);return;}
    startPolling();
  });
}
function startPolling(){
  if(pollTimer)clearInterval(pollTimer);
  pollTimer=setInterval(pollStatus,800);
  pollStatus();
}
function pollStatus(){
  fetch('/api/scrape/status').then(r=>r.json()).then(d=>{
    const running=d.running;
    const result=d.result;
    // Button
    const btn=$('run-btn');
    if(running){
      btn.disabled=true;btn.classList.add('running');
      $('run-icon').textContent='⏳';$('run-label').textContent='Running...';
    }else{
      btn.disabled=false;btn.classList.remove('running');
      $('run-icon').textContent='▶';$('run-label').textContent='Start Scraping';
    }
    // Badge
    const badge=$('status-badge');const dot=$('pulse-dot');const stxt=$('status-text');
    badge.className='status-badge '+(running?'running':result||'idle');
    dot.className='pulse'+(running?' anim':'');
    stxt.textContent=running?'Running...':(result==='ok'?'Done ✓':result==='error'?'Error':'Idle');
    // Times
    if(d.started||d.finished){
      $('run-times').style.display='block';
      $('run-started').textContent=d.started?'Started: '+d.started.replace('T',' ').slice(0,19):'';
      $('run-finished').textContent=d.finished?'Finished: '+d.finished.replace('T',' ').slice(0,19):'';
    }
    // Log
    const lb=$('log-box');
    if(d.log&&d.log.length){
      lb.innerHTML=d.log.map(colorLine).join('\n');
      lb.scrollTop=lb.scrollHeight;
    }
    if(!running&&pollTimer){clearInterval(pollTimer);pollTimer=null;loadStats();loadNavStats();metaLoaded=false;}
  });
}
function colorLine(l){
  const e=esc(l);
  if(l.includes('[OK]')||l.includes('[JOB]')||l.includes('Logged in'))return`<span class="log-ok">${e}</span>`;
  if(l.includes('[JOB]'))return`<span class="log-job">${e}</span>`;
  if(l.includes('[SKIP-SPAM]'))return`<span class="log-spam">${e}</span>`;
  if(l.includes('[WARN]')||l.includes('WARN'))return`<span class="log-warn">${e}</span>`;
  if(l.includes('ERROR')||l.includes('[ERR]'))return`<span class="log-err">${e}</span>`;
  if(l.includes('[STEP]')||l.includes('SEARCH INPUT')||l.includes('MATCH'))return`<span class="log-step">${e}</span>`;
  return`<span class="log-info">${e}</span>`;
}
// Check if already running on page load
pollStatus();

// ══════════════════════════════════════════════════════════════
//  JOBS
// ══════════════════════════════════════════════════════════════
const togState={has_rate:false,has_email:false};
function toggleTog(btn){
  const f=btn.dataset.f;togState[f]=!togState[f];
  btn.classList.toggle('active',togState[f]);loadJobs();
}
function resetJobFilters(){
  $('jtitle').value='';$('jskill').value='';$('jloc').value='';
  ['jcontract','jvisa','jgroup'].forEach(id=>$(id).value='');
  document.querySelectorAll('#jexp-checks input').forEach(c=>c.checked=false);
  Object.keys(togState).forEach(k=>togState[k]=false);
  document.querySelectorAll('.tog-btn').forEach(b=>b.classList.remove('active'));
  loadJobs();
}

let metaLoaded=false;
function loadJobMeta(){
  // Only load once — don't reset dropdowns on every tab switch
  if(metaLoaded)return;
  fetch('/api/meta').then(r=>r.json()).then(d=>{
    metaLoaded=true;
    const cs=$('jcontract');cs.innerHTML='<option value="">All types</option>';
    d.contracts.forEach(c=>{const o=document.createElement('option');o.value=c;o.textContent=c;cs.appendChild(o);});
    const vs=$('jvisa');vs.innerHTML='<option value="">All visas</option>';
    d.visas.forEach(v=>{const o=document.createElement('option');o.value=v;o.textContent=v;vs.appendChild(o);});
    const gs=$('jgroup');gs.innerHTML='<option value="">All groups</option>';
    d.groups.forEach(g=>{const o=document.createElement('option');o.value=g;o.textContent=g;gs.appendChild(o);});
    const ec=$('jexp-checks');ec.innerHTML='';
    (d.exp_buckets||[]).forEach(b=>{
      const l=document.createElement('label');const cb=document.createElement('input');
      cb.type='checkbox';cb.value=b;cb.addEventListener('change',loadJobs);
      l.appendChild(cb);l.appendChild(document.createTextNode(b+' yrs'));ec.appendChild(l);
    });
    // Attach select listeners after meta is loaded
    ['jcontract','jvisa','jgroup'].forEach(id=>{
      $(id).addEventListener('change',loadJobs);
    });
    loadJobs();
  });
}

let jdeb;
function loadJobs(){
  const p=new URLSearchParams();
  const tv=$('jtitle').value.trim();if(tv)p.set('title',tv);
  const sv=$('jskill').value.trim();if(sv)p.set('skill',sv);
  const cv=$('jcontract').value;if(cv)p.set('contract',cv);
  const vv=$('jvisa').value;if(vv)p.set('visa',vv);
  const gv=$('jgroup').value;if(gv)p.set('group',gv);
  const lv=$('jloc').value.trim();if(lv)p.set('location',lv);
  const eb=[...$('jexp-checks').querySelectorAll('input:checked')].map(c=>c.value);
  if(eb.length)p.set('experience',eb.join(','));
  if(togState.has_rate)p.set('has_rate','1');
  if(togState.has_email)p.set('has_email','1');
  fetch('/api/jobs?'+p).then(r=>r.json()).then(d=>{
    $('jobs-stats').innerHTML=`Showing <b>${d.filtered}</b> of <b>${d.total}</b> jobs`;
    renderJobs(d.jobs);
  });
}
function renderJobs(jobs){
  const tb=$('jobs-rows');const em=$('jobs-empty');
  tb.innerHTML='';
  if(!jobs.length){em.style.display='block';return;}em.style.display='none';
  jobs.forEach(j=>{
    const tr=document.createElement('tr');tr.onclick=()=>openModal(j);
    let pills='';
    if(j.contract_type)pills+=ctag(j.contract_type);
    (j.visa_types||[]).slice(0,3).forEach(v=>pills+=tag(v,'tvisa'));
    let skills='';(j.skills||[]).slice(0,3).forEach(s=>skills+=tag(s,'tskill'));
    let contact='';
    if(j.contact_email)contact+=`<a href="mailto:${esc(j.contact_email)}" onclick="event.stopPropagation()">${esc(j.contact_email)}</a>`;
    if(j.contact_phone)contact+=`<span style="display:block;color:var(--dim)">${esc(j.contact_phone)}</span>`;
    if(!contact)contact=`<span style="color:var(--border)">—</span>`;
    tr.innerHTML=`
      <td><div class="cell-title">${esc(j.job_title||'–')}</div>
          <div class="cell-group">${esc(j.source_group||'')}</div>
          <div class="cell-tags">${pills}${skills}</div></td>
      <td class="mono">${esc(j.location||'–')}</td>
      <td class="mono">${esc(j.experience||'–')}</td>
      <td>${normalizeRateDisplay(j.rate)}</td>
      <td class="mono">${esc(j.duration||'–')}</td>
      <td><div class="cell-tags">${ctag(j.contract_type)}${(j.visa_types||[]).slice(0,2).map(v=>tag(v,'tvisa')).join('')}</div></td>
      <td class="contact-cell">${contact}</td>
      <td class="time-cell">${timeAgo(j.wa_timestamp||j.scraped_at)}</td>`;
    tb.appendChild(tr);
  });
}

// ══════════════════════════════════════════════════════════════
//  REVIEW
// ══════════════════════════════════════════════════════════════
let reviewTab='all';
function setReviewTab(tab,el){
  reviewTab=tab;
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  el.classList.add('active');loadReview();
}
function loadReview(){
  const p=new URLSearchParams();
  if(reviewTab!=='all')p.set('status',reviewTab);
  fetch('/api/review?'+p).then(r=>r.json()).then(d=>{
    $('review-stats').innerHTML=`Showing <b>${d.filtered}</b> of <b>${d.total}</b> messages`;
    renderReview(d.data);
  });
}
function renderReview(rows){
  const tb=$('review-rows');const em=$('review-empty');
  tb.innerHTML='';
  if(!rows.length){em.style.display='block';return;}em.style.display='none';
  rows.forEach(r=>{
    const tr=document.createElement('tr');tr.onclick=()=>openReviewModal(r);
    const stag=r.status==='job'?tag('job','tok'):r.status==='spam'?tag('spam','tspam'):tag('no_kw','tvisa');
    tr.innerHTML=`
      <td>${stag}</td>
      <td class="mono">${esc((r.source_group||'').replace('| C2C | REQUIREMENTS |','|C2C| REQ').slice(0,22))}</td>
      <td class="time-cell">${timeAgo(r.wa_timestamp)}</td>
      <td class="mono" style="font-size:11px">${esc((r.filter_reason||'').slice(0,60))}</td>
      <td class="mono" style="font-size:11px;white-space:pre-wrap">${esc((r.raw_message||'').slice(0,120))}${r.raw_message&&r.raw_message.length>120?'…':''}</td>`;
    tb.appendChild(tr);
  });
}

// ══════════════════════════════════════════════════════════════
//  MODAL
// ══════════════════════════════════════════════════════════════
function mf(label,val,cls='',full=false){
  if(!val&&val!==0)return'';
  return`<div class="mf${full?' full':''}"><div class="ml">${label}</div><div class="mv ${cls}">${val}</div></div>`;
}
function openModal(j){
  $('m-title').textContent=j.job_title||'Untitled';
  let pills='';
  if(j.contract_type)pills+=ctag(j.contract_type);
  (j.visa_types||[]).forEach(v=>pills+=tag(v,'tvisa'));
  (j.skills||[]).forEach(s=>pills+=tag(s,'tskill'));
  $('m-pills').innerHTML=pills;
  const emailLink=j.contact_email?`<a href="mailto:${esc(j.contact_email)}">${esc(j.contact_email)}</a>`:null;
  const phone=j.contact_phone?`<a href="tel:${esc(j.contact_phone)}">${esc(j.contact_phone)}</a>`:null;
  $('m-grid').innerHTML=[
    mf('Location',esc(j.location)),
    mf('Experience',esc(j.experience)),
    mf('Rate',esc(j.rate),'gv'),
    mf('Duration',esc(j.duration)),
    mf('Contract Type',esc(j.contract_type)),
    mf('Client',esc(j.client)),
    mf('Vendor',esc(j.vendor)),
    mf('Interview',esc(j.interview_type)),
    mf('Contact',[emailLink,phone].filter(Boolean).join('<br>')||null),
    mf('Contact Name',esc(j.contact_name)),
    mf('Apply Link',j.apply_link?`<a href="${esc(j.apply_link)}" target="_blank">${esc(j.apply_link)}</a>`:null),
    mf('Source Group',esc(j.source_group)),
    mf('Sender',esc(j.sender)),
    mf('Posted',esc(j.wa_timestamp)),
  ].join('');
  $('m-raw').textContent=j.raw_message||'';
  $('m-raw').style.display='none';
  $('raw-toggle').textContent='Show raw message';
  $('overlay').classList.add('open');
}
function openReviewModal(r){
  $('m-title').textContent=r.status==='job'?'Message: Job':'Message: '+(r.status||'unknown');
  $('m-pills').innerHTML=(r.status==='job'?tag('job','tok'):r.status==='spam'?tag('spam','tspam'):tag('no_kw','tvisa'))+
    (r.source_group?tag(r.source_group.slice(0,20),'tvisa'):'');
  $('m-grid').innerHTML=[
    mf('Status',esc(r.status)),
    mf('Filter Reason',esc(r.filter_reason)),
    mf('Source Group',esc(r.source_group)),
    mf('Sender',esc(r.sender)),
    mf('Posted',esc(r.wa_timestamp)),
    mf('Scraped At',esc(r.scraped_at)),
  ].join('');
  $('m-raw').textContent=r.raw_message||'';
  $('m-raw').style.display='block';
  $('raw-toggle').textContent='Hide raw message';
  $('overlay').classList.add('open');
}
function closeModal(){$('overlay').classList.remove('open')}
function toggleRaw(){
  const r=$('m-raw'),cb=$('copy-raw-btn'),s=r.style.display!=='none';
  r.style.display=s?'none':'block';
  if(cb)cb.style.display=s?'none':'inline-block';
  $('raw-toggle').textContent=s?'Show raw message':'Hide raw message';
}
function copyRaw(){
  const text=$('m-raw').textContent;
  navigator.clipboard.writeText(text).then(()=>{
    const b=$('copy-raw-btn'),o=b.textContent;
    b.textContent='✓ Copied!';b.style.color='var(--accent)';b.style.borderColor='var(--accent)';
    setTimeout(()=>{b.textContent=o;b.style.color='';b.style.borderColor='';},2000);
  }).catch(()=>{
    const ta=document.createElement('textarea');ta.value=$('m-raw').textContent;
    document.body.appendChild(ta);ta.select();document.execCommand('copy');document.body.removeChild(ta);
    const b=$('copy-raw-btn');b.textContent='✓ Copied!';setTimeout(()=>b.textContent='📋 Copy',2000);
  });
}
$('overlay').onclick=e=>{if(e.target.id==='overlay')closeModal();}
document.addEventListener('keydown',e=>{if(e.key==='Escape')closeModal();});

// ── Theme toggle ────────────────────────────────────────────────────────
function toggleTheme(){
  const isLight=document.body.classList.toggle('light');
  const btn=$('theme-btn');
  btn.textContent=isLight?'☀️ Light':'🌙 Dark';
  localStorage.setItem('theme',isLight?'light':'dark');
}
(function(){
  if(localStorage.getItem('theme')==='light'){
    document.body.classList.add('light');
    const b=$('theme-btn');if(b)b.textContent='☀️ Light';
  }
})();

// debounce text inputs
let dTimer;
['jtitle','jskill','jloc'].forEach(id=>{
  $(id).addEventListener('input',()=>{clearTimeout(dTimer);dTimer=setTimeout(loadJobs,280);});
});

// ── Excel export ───────────────────────────────────────────────────────────
function exportExcel(){
  const btn=event.target;
  btn.textContent='⏳ Exporting...';btn.disabled=true;
  fetch('/api/export/excel')
    .then(r=>{
      if(!r.ok)return r.json().then(d=>{throw new Error(d.error||'Export failed');});
      return r.blob();
    })
    .then(blob=>{
      const url=URL.createObjectURL(blob);
      const a=document.createElement('a');
      const now=new Date();
      const dt=`${now.getFullYear()}${String(now.getMonth()+1).padStart(2,'0')}${String(now.getDate()).padStart(2,'0')}`;
      a.href=url;a.download=`c2c_jobs_${dt}.xlsx`;
      document.body.appendChild(a);a.click();
      setTimeout(()=>{URL.revokeObjectURL(url);a.remove();},1000);
    })
    .catch(err=>alert('Export error: '+err.message))
    .finally(()=>{btn.textContent='⬇ Export Excel';btn.disabled=false;});
}

// ── Rate normalization display ─────────────────────────────────────────────
function normalizeRateDisplay(rate){
  if(!rate)return'<span style="color:var(--border)">—</span>';
  // Extract numbers that look like $/hr amounts (10-500 range)
  const nums=[...rate.matchAll(/\$?\s*(\d+(?:\.\d+)?)/g)]
    .map(m=>parseFloat(m[1]))
    .filter(n=>n>=10&&n<=500);
  if(!nums.length)return`<span class="green-val">${esc(rate)}</span>`;
  const lo=Math.min(...nums),hi=Math.max(...nums);
  const display=lo===hi?`$${lo}/hr`:`$${lo}–$${hi}/hr`;
  return`<span class="green-val" title="${esc(rate)}">${display}</span>`;
}
</script>
</body>
</html>"""

if __name__=="__main__":
    import os
    if len(sys.argv)>1:
        DATA_PATH=Path(sys.argv[1]).resolve()
    # Render and other cloud hosts set PORT env variable
    port = int(os.environ.get("PORT", 5000))
    # On cloud: bind to 0.0.0.0 so it's publicly accessible
    # On local: bind to 127.0.0.1 for security
    host = "0.0.0.0" if os.environ.get("PORT") else "127.0.0.1"
    print("="*55)
    print("  C2C WhatsApp Job Scraper — Unified App")
    print(f"  Folder : {SDIR}")
    print(f"  Output : {OUT}")
    print(f"  URL    : http://{host}:{port}")
    print("="*55)
    app.run(host=host, port=port, debug=False)

    #done